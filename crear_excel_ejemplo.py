#!/usr/bin/env python3
"""
Script para crear un archivo Excel de ejemplo para importar alumnos
"""

try:
    from openpyxl import Workbook

    # Crear un nuevo libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Alumnos"

    # Formato 1: Nombre(s) - Apellidos
    ws['A1'] = "Nombre Completo"
    ws['A2'] = "Juan García López"
    ws['A3'] = "María Fernanda Rodríguez Martínez"
    ws['A4'] = "Carlos Hernández Pérez"
    ws['A5'] = "Ana Sofía González Ramírez"
    ws['A6'] = "Luis Miguel Torres Sánchez"

    # Guardar archivo
    wb.save("ejemplo_alumnos_formato_nombre_apellidos.xlsx")
    print("✓ Archivo creado: ejemplo_alumnos_formato_nombre_apellidos.xlsx")

    # Crear otro archivo con formato inverso
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Alumnos"

    # Formato 2: Apellidos - Nombre(s)
    ws2['A1'] = "Nombre Completo"
    ws2['A2'] = "García López Juan"
    ws2['A3'] = "Rodríguez Martínez María Fernanda"
    ws2['A4'] = "Hernández Pérez Carlos"
    ws2['A5'] = "González Ramírez Ana Sofía"
    ws2['A6'] = "Torres Sánchez Luis Miguel"

    # Guardar archivo
    wb2.save("ejemplo_alumnos_formato_apellidos_nombre.xlsx")
    print("✓ Archivo creado: ejemplo_alumnos_formato_apellidos_nombre.xlsx")

    print("\n✅ Archivos de ejemplo creados exitosamente")
    print("\nFormato 1 (Nombre-Apellidos):")
    print("  - Juan García López")
    print("  - María Fernanda Rodríguez Martínez")
    print("\nFormato 2 (Apellidos-Nombre):")
    print("  - García López Juan")
    print("  - Rodríguez Martínez María Fernanda")

except ImportError:
    print("❌ Error: Se requiere la librería openpyxl")
    print("Instalar con: pip install openpyxl")
except Exception as e:
    print(f"❌ Error: {e}")
